"""Export Brandýs shipping data from Vnitro_Brandys_final.xlsx to JSON."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

APP_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = Path(r"e:\AI\vnitro dopravy\Brandys\Vnitro_Brandys_final.xlsx")
DEFAULT_OUT = APP_DIR / "data" / "shipping.json"

VNITRO_CARRIERS = ["FLEXILOG", "MB Doprava", "LEŠTINA", "LogEx", "Q CARGO"]
PRAHA_CARRIERS = ["FLEXILOG", "MB Doprava", "Q CARGO"]

WEIGHT_CATEGORIES = [
    {"key": "2000", "label": "do 2 tun", "max_kg": 2000},
    {"key": "3000", "label": "do 3 tun", "max_kg": 3000},
    {"key": "6000", "label": "do 6 tun", "max_kg": 6000},
    {"key": "8000", "label": "do 8 tun", "max_kg": 8000},
    {"key": "12000", "label": "do 12 tun", "max_kg": 12000},
    {"key": "999999", "label": "nad 12 tun", "max_kg": 999999},
]

PRAHA_WEIGHT_CATEGORIES = [
    {"key": "2000", "label": "do 2 tun", "max_kg": 2000},
    {"key": "3500", "label": "do 3,5 tun", "max_kg": 3500},
    {"key": "6000", "label": "do 6 tun", "max_kg": 6000},
    {"key": "8000", "label": "do 8 tun", "max_kg": 8000},
    {"key": "12000", "label": "do 12 tun", "max_kg": 12000},
    {"key": "999999", "label": "nad 12 tun", "max_kg": 999999},
]


def as_float(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None


def normalize_carrier(name: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(name).strip())
    replacements = {
        "MB  Doprava": "MB Doprava",
        "FLEXILOG (do 3,5t)": "FLEXILOG",
        "Q CARGO (do 3 tun)": "Q CARGO",
        "Q CARGO (do 11t)": "Q CARGO",
        "Q CARGO (nad 11t)": "Q CARGO",
        "Q CARGO ": "Q CARGO",
    }
    return replacements.get(cleaned, cleaned)


def parse_zony_brandys(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb["Zóny Brandýs"]
    entries: list[dict] = []
    for row in range(2, ws.max_row + 1):
        psc_od = ws.cell(row, 1).value
        psc_do = ws.cell(row, 2).value
        zone = ws.cell(row, 3).value
        city = ws.cell(row, 5).value
        if psc_od is None or psc_do is None or zone is None:
            continue
        zone_str = str(int(zone)) if isinstance(zone, (int, float)) else str(zone).strip()
        entries.append(
            {
                "psc_od": int(psc_od),
                "psc_do": int(psc_do),
                "zone": zone_str,
                "city": str(city) if city else "",
            }
        )
    return entries


def parse_raben(wb: openpyxl.Workbook) -> dict:
    ws = wb["Raben_Brandys"]
    weights: list[int] = []
    prices: dict[str, dict[str, float | None]] = {str(zone): {} for zone in range(1, 9)}

    for row in range(14, ws.max_row + 1):
        weight = ws.cell(row, 3).value
        if not isinstance(weight, (int, float)):
            break
        weight_int = int(weight)
        if weight_int in weights:
            continue
        row_prices: dict[str, float | None] = {}
        for zone_index, col in enumerate(range(5, 13), start=1):
            row_prices[str(zone_index)] = as_float(ws.cell(row, col).value)
        if any(value is not None for value in row_prices.values()):
            weights.append(weight_int)
            for zone_key, price in row_prices.items():
                prices[zone_key][str(weight_int)] = price

    weights.sort()
    psc_ranges: list[dict] = []
    for row in range(23, 40):
        label = ws.cell(row, 15).value
        psc_od = ws.cell(row, 16).value
        psc_do = ws.cell(row, 17).value
        if not isinstance(psc_od, (int, float)) or not isinstance(psc_do, (int, float)):
            continue
        match = re.search(r"(\d+)", str(label or ""))
        if not match:
            continue
        psc_ranges.append(
            {
                "psc_od": int(psc_od),
                "psc_do": int(psc_do),
                "zone": int(match.group(1)),
            }
        )

    return {"psc_ranges": psc_ranges, "weights": weights, "prices": prices, "max_weight_kg": weights[-1] if weights else None}


def parse_dnp(wb: openpyxl.Workbook) -> dict:
    ws = wb["DNP_Brandys"]
    weights: list[int] = []
    for col in range(2, ws.max_column + 1):
        weight = ws.cell(2, col).value
        if isinstance(weight, (int, float)):
            weights.append(int(weight))
        elif weights:
            break

    prices: dict[str, dict[str, float | None]] = {}
    for row in range(3, ws.max_row + 1):
        zone = ws.cell(row, 1).value
        if not isinstance(zone, (int, float)) or zone > 100:
            break
        zone_key = str(int(zone))
        prices[zone_key] = {}
        for col, weight in enumerate(weights, start=2):
            prices[zone_key][str(weight)] = as_float(ws.cell(row, col).value)

    psc_ranges: list[dict] = []
    for row in range(23, ws.max_row + 1):
        psc_do = ws.cell(row, 1).value
        psc_od = ws.cell(row, 2).value
        zone = ws.cell(row, 5).value
        if not isinstance(psc_od, (int, float)) or not isinstance(psc_do, (int, float)):
            continue
        if not isinstance(zone, (int, float)):
            continue
        psc_ranges.append(
            {
                "psc_od": int(psc_od),
                "psc_do": int(psc_do),
                "zone": int(zone),
                "okres": str(ws.cell(row, 4).value or ""),
            }
        )

    return {"psc_ranges": psc_ranges, "weights": weights, "prices": prices, "max_weight_kg": weights[-1] if weights else None}


def read_vnitro_category_block(
    ws: openpyxl.Workbook.worksheets,
    row: int,
    start_col: int,
) -> dict:
    limit_cn = as_float(ws.cell(row, start_col).value)
    minimum = as_float(ws.cell(row, start_col + 1).value)
    carriers: dict[str, float | None] = {}
    for offset, carrier in enumerate(VNITRO_CARRIERS, start=3):
        carriers[carrier] = as_float(ws.cell(row, start_col + offset).value)
    return {"limit_cn": limit_cn, "minimum": minimum, "carriers": carriers}


def parse_vnitro(wb: openpyxl.Workbook) -> dict:
    ws = wb["Vnitro_Brandys"]
    fuel_surcharge = as_float(ws.cell(4, 4).value) or 0.12
    additional_unload = as_float(ws.cell(9, 6).value) or 500.0

    light_blocks = [
        ("2000", 3),
        ("3000", 11),
        ("6000", 19),
    ]
    heavy_blocks = [
        ("8000", 3),
        ("12000", 11),
        ("999999", 19),
    ]

    zones: dict[str, dict] = {}
    for row in range(10, 33):
        zone = ws.cell(row, 1).value
        if not isinstance(zone, (int, float)):
            continue
        zone_key = str(int(zone))
        zone_data: dict[str, object] = {"km_range": str(ws.cell(row, 2).value or "")}
        for weight_key, start_col in light_blocks:
            zone_data[weight_key] = read_vnitro_category_block(ws, row, start_col)
        zones[zone_key] = zone_data

    for row in range(37, 70):
        zone = ws.cell(row, 1).value
        if not isinstance(zone, (int, float)):
            if zones and row > 37:
                break
            continue
        zone_key = str(int(zone))
        if zone_key not in zones:
            zones[zone_key] = {"km_range": str(ws.cell(row, 2).value or "")}
        for weight_key, start_col in heavy_blocks:
            zones[zone_key][weight_key] = read_vnitro_category_block(ws, row, start_col)

    return {
        "fuel_surcharge": fuel_surcharge,
        "additional_unload": additional_unload,
        "weight_categories": WEIGHT_CATEGORIES,
        "zones": zones,
    }


def read_praha_block(ws: openpyxl.Workbook.worksheets, row: int, start_col: int) -> dict[str, float | None]:
    carriers: dict[str, float | None] = {}
    for offset, carrier in enumerate(PRAHA_CARRIERS):
        carriers[carrier] = as_float(ws.cell(row, start_col + offset).value)
    return carriers


def parse_praha(wb: openpyxl.Workbook) -> dict:
    ws = wb["Zony 1P-6P"]
    additional_unload = as_float(ws.cell(11, 2).value) or 500.0
    zones: dict[str, dict[str, dict[str, float | None]]] = {}

    light_layout = [
        ("2000", 2),
        ("3500", 5),
        ("6000", 8),
    ]
    heavy_layout = [
        ("8000", 2),
        ("12000", 5),
        ("999999", 8),
    ]

    for row in range(12, 18):
        zone = ws.cell(row, 1).value
        if not zone:
            continue
        zone_key = str(zone).strip()
        zones[zone_key] = {}
        for weight_key, start_col in light_layout:
            zones[zone_key][weight_key] = read_praha_block(ws, row, start_col)

    for row in range(23, 29):
        zone = ws.cell(row, 1).value
        if not zone:
            continue
        zone_key = str(zone).strip()
        if zone_key not in zones:
            zones[zone_key] = {}
        for weight_key, start_col in heavy_layout:
            zones[zone_key][weight_key] = read_praha_block(ws, row, start_col)

    return {
        "additional_unload": additional_unload,
        "weight_categories": PRAHA_WEIGHT_CATEGORIES,
        "zones": zones,
    }


def build_data(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    return {
        "meta": {
            "source": xlsx_path.name,
            "branch": "Brandýs nad Labem",
            "carriers": {
                "pallet": ["Raben", "DNP"],
                "vnitro": VNITRO_CARRIERS,
                "praha": PRAHA_CARRIERS,
            },
        },
        "zony": parse_zony_brandys(wb),
        "raben": parse_raben(wb),
        "dnp": parse_dnp(wb),
        "vnitro": parse_vnitro(wb),
        "praha": parse_praha(wb),
    }


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT

    if not xlsx_path.exists():
        print(f"Chyba: soubor {xlsx_path} neexistuje.", file=sys.stderr)
        sys.exit(1)

    data = build_data(xlsx_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Export dokoncen: {out_path}")
    print(f"  Zony Brandys: {len(data['zony'])} intervalu PSC")
    print(f"  Raben: {len(data['raben']['weights'])} hmotnostnich pasem")
    print(f"  DNP: {len(data['dnp']['prices'])} zon")
    print(f"  Vnitro: {len(data['vnitro']['zones'])} pasem")
    print(f"  Praha P: {len(data['praha']['zones'])} pasem")


if __name__ == "__main__":
    main()
